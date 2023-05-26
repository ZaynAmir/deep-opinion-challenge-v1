from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
from django.db import transaction
from django.http import HttpResponse, Http404
from django.core.exceptions import ObjectDoesNotExist
import pandas as pd
import os
import openpyxl
import sqlite3
import redis
from .models import SheetModel, TrainingData, Tag
from .serializers import TagSerializer


# Establish a Redis connection
redis_connection = redis.Redis(host='localhost', port=6379, db=0)


class FileUploadView(APIView):
    def post(self, request, format=None):
        file = request.FILES['file']

        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.name.endswith('.xlsx'):
            df = pd.read_excel(file)
        else:
            return Response({'error': 'Invalid file format'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            column = df.columns.to_list()[0]
            df_to_dict = df.to_dict('records')
            file_name = os.path.basename(file.name)
            rows_count = len(df_to_dict)

            sheet = SheetModel.objects.create(
                name = file_name,
                row_count = rows_count
            )
            conn = sqlite3.connect('db.sqlite3')

            # Increase the cache size
            conn.execute("PRAGMA cache_size = 6000000")  # 3GB cache size

            # Set journaling mode to memory
            conn.execute("PRAGMA journal_mode = MEMORY")

            # Disable indexing while inserting
            conn.execute("PRAGMA defer_foreign_keys = ON")
            conn.execute("PRAGMA synchronous = OFF")
            conn.execute("PRAGMA count_changes = OFF")
            conn.execute("PRAGMA temp_store = MEMORY")

            # Create a cursor object
            cursor = conn.cursor()

            sql = "INSERT INTO app_trainingdata (text, sheet_id) VALUES (?, ?)"
            # Prepare the data for insertion (a list of tuples)
            data = [(item[column], sheet.id) for item in df_to_dict if item[column] is not None and str(item[column]) != 'nan']

            # Execute the bulk insert
            cursor.executemany(sql, data)

            # Commit the changes
            conn.commit()

            # Close the cursor and the connection
            cursor.close()
            conn.close()
            return Response({"success": True, "sheet_id":sheet.id, "message": f"{file_name} uploaded."} ,status=status.HTTP_201_CREATED)
        except Exception as e:
            print(e)
            return Response({'error': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class CreateTag(APIView):
    def post(self, request, *args, **kwargs):
            try:
                aspect = request.data.get("aspect")
                sentiment = request.data.get("sentiment")
                data_id = request.data.get("text_id")
                if not aspect or not sentiment or not data_id:
                    hint = {
                        "text_id": "required",
                        "sentiment": "required",
                        "aspect": "required"
                    }
                    return Response({"success": False, 'error': 'request payload is not complete', "hint": hint }, status=status.HTTP_400_BAD_REQUEST)
                if sentiment not in ["POS", "NEG", "NEU"]:
                    return Response({"success": False, 'error': 'sentiment is not in ("POS", "NEG", "NEU")'}, status=status.HTTP_400_BAD_REQUEST)
                with transaction.atomic(): # applying rollback transaction to make sure data integrity
                    new_tag = Tag.objects.create(aspect=aspect, sentiment=sentiment)
                    training_data =  TrainingData.objects.get(id=data_id)
                    training_data.tags.add(new_tag)
                    training_data.save()
                return Response({"success": True, "tag_id": new_tag.id,  "message": "tag is created" } ,status=status.HTTP_201_CREATED)
            except TrainingData.DoesNotExist:
                return Response({"success": False, 'error': 'text_id on which you are trying to create tag is not avaliable'}, status=status.HTTP_400_BAD_REQUEST)
            except:
                return Response({"success": False, 'error': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
class UpdateTag(APIView):
    def put(self, request, *args, **kwargs):
            try:
                tag_id = kwargs["tag_id"]
                tag_instance =  Tag.objects.get(id=tag_id)
                serializer = TagSerializer(instance=tag_instance, data=request.data ,partial=True)
                if serializer.is_valid(raise_exception=True):
                    serializer.save()
                    return Response({"success": True,  "message": "tag is updated", "data":serializer.data } ,status=status.HTTP_200_OK)
            except Tag.DoesNotExist:
                return Response({"success": False, 'error': 'tag which you are trying to update is not avaliable'}, status=status.HTTP_400_BAD_REQUEST)
            

class GetAvaliableAspects(APIView):
    def get(self, request, *args, **kwargs):
        sheet_id = kwargs["sheet_id"]
        try:
            sheet = SheetModel.objects.get(id=sheet_id)
        except SheetModel.DoesNotExist:
            return Response({"success": False, 'error': f'sheet with id: {sheet_id} is not avaliable'}, status=status.HTTP_400_BAD_REQUEST)
        queryset = Tag.objects.filter(trainingdata__sheet=sheet).values("aspect").distinct().order_by("aspect")
        paginator = PageNumberPagination()
        paginator.page_size = request.query_params.get('page_size', 10)  # Number of items per page
        paginated_queryset = paginator.paginate_queryset(queryset, request)
        return paginator.get_paginated_response(paginated_queryset)
    
class GetAvaliableSentiments(APIView):
    def get(self, request, *args, **kwargs):
        sheet_id = kwargs["sheet_id"]
        try:
            sheet = SheetModel.objects.get(id=sheet_id)
        except SheetModel.DoesNotExist:
            return Response({"success": False, 'error': f'sheet with id: {sheet_id} is not avaliable'}, status=status.HTTP_400_BAD_REQUEST)
        # queryset = Tag.objects.all().values("sentiment").distinct()
        queryset = Tag.objects.filter(trainingdata__sheet=sheet).values("sentiment").distinct().order_by("sentiment")
        paginator = PageNumberPagination()
        paginator.page_size = request.query_params.get('page_size', 10)  # Number of items per page
        paginated_queryset = paginator.paginate_queryset(queryset, request)
        return paginator.get_paginated_response(paginated_queryset)
    

class GetTextDataWithTags(APIView):
    def get(self, request, *args, **kwargs):
        document_id = kwargs["sheet_id"]
        if SheetModel.objects.filter(id=document_id).exists():
            queryset = TrainingData.objects.filter(sheet=document_id)
            paginator = PageNumberPagination()
            paginator.page_size = request.query_params.get('page_size', 10)  # Number of items per page
            paginated_queryset = paginator.paginate_queryset(queryset, request)
            data = [{"id": item.id, "text": item.text, "tags": TagSerializer(instance=item.tags.all(), many=True).data} for item in paginated_queryset]
            return paginator.get_paginated_response(data)

        return Response({"success": False, "error": f"Sheet with id: {document_id} is not found"}, status=status.HTTP_404_NOT_FOUND)
    


# utility function to save excel sheet in xlsx format
def save_data_to_xlsx(data, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Write headers
    sheet['A1'] = 'text'
    sheet['B1'] = 'aspect'
    sheet['C1'] = 'sentiment'
    # Write data
    row = 2  # Start from the second row
    for entry in data:
        text = entry['text']
        tags = entry['tags']
        for tag in tags:
            aspect = tag['aspect']
            sentiment = tag['sentiment']
            sheet.cell(row=row, column=1, value=text)
            sheet.cell(row=row, column=2, value=aspect)
            sheet.cell(row=row, column=3, value=sentiment)
            row += 1
    # Save the workbook
    workbook.save(file_path)


# def download_sheet(request, sheet_id):
#     try:
#         sheet = TrainingData.objects.filter(sheet=sheet_id)
#         if sheet:
#             data = [{"text": item.text, "tags": TagSerializer(instance=item.tags.all(), many=True).data} for item in sheet]
#             file_name_with_extension = SheetModel.objects.filter(id=sheet_id).first()
#             file_name = file_name_with_extension.name.split(".")[0]
#             file_path = f'{file_name}.xlsx'
#             save_data_to_xlsx(data, file_path)

#             with open(file_path, 'rb') as file:
#                 response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#                 response['Content-Disposition'] = f'attachment; filename={file_name}-tags.xlsx'
#             return response
#         else:
#             raise Http404("Sheet not found.")
#     except ObjectDoesNotExist:
#         raise Http404("Sheet not found.")

def download_sheet(request, sheet_id):
    try:
        # Check if the file exists in Redis cache
        cached_file = redis_connection.get(f'sheet:{sheet_id}')
        if cached_file:
            # If the file exists in cache, return it as a response
            file_name_with_extension = SheetModel.objects.filter(id=sheet_id).first()
            file_name = file_name_with_extension.name.split(".")[0]
            response = HttpResponse(cached_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={file_name}-tags.xlsx'
            return response

        # If the file doesn't exist in cache, generate it and save it to cache
        sheet = TrainingData.objects.filter(sheet=sheet_id)
        if sheet:
            data = [{"text": item.text, "tags": TagSerializer(instance=item.tags.all(), many=True).data} for item in sheet]
            file_name_with_extension = SheetModel.objects.filter(id=sheet_id).first()
            file_name = file_name_with_extension.name.split(".")[0]
            file_path = f'{file_name}.xlsx'
            save_data_to_xlsx(data, file_path)

            with open(file_path, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={file_name}-tags.xlsx'

                # Save the file to Redis cache for future requests
                redis_connection.set(f'sheet:{sheet_id}', response.content)

                # Set expiration time for the cached file (e.g., 1 hour)
                redis_connection.expire(f'sheet:{sheet_id}', 3600)

            return response
        else:
            raise Http404("Sheet not found.")
    except ObjectDoesNotExist:
        raise Http404("Sheet not found.")